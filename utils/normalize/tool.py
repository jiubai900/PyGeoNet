from openpyxl.workbook import Workbook
from typing import Union, List
from datetime import datetime
import re
from io import StringIO
import gzip
import tarfile
import time
from GEOparse import get_GEO
from bs4 import BeautifulSoup
import requests
import tqdm
import logging
import pandas as pd
import os
import shutil
from fuzzywuzzy import fuzz
from collections import OrderedDict

sep = ['\t', ';', ',', ' ']  # Separators for parsing CSV files

NULL_VALUE = 'null'
GPL_ID_SEPARATOR = ' '
logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)
# Check if FileHandler already exists to avoid duplicate addition
if not any(
        isinstance(handler, logging.FileHandler) and handler.baseFilename == 'app.log' for handler in logger.handlers):
    # If not, create and add
    file_handler = logging.FileHandler('app.log')
    formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
    file_handler.setFormatter(formatter)
    logger.addHandler(file_handler)


def analyse_matrix(matrix_path, gpl_path, suppl_path, save_path):
    gse_number = os.path.basename(matrix_path)
    if not os.path.exists(save_path):
        os.makedirs(save_path)
    matrix_path = os.path.join(matrix_path, 'matrix')
    if 'GSE' == gse_number[0:3]:
        file_arr = []
        for root, dirs, files in os.walk(matrix_path):
            for file in files:
                file_arr.append(os.path.join(root, file))
        get_file(matrix_path, file_arr, gse_number, gpl_path, suppl_path, save_path)
    else:
        logger.error(f"{matrix_path} is not a valid directory.")


def get_file(matrix_path, files, number, gpl_path, suppl_path, save_path):
    """
    Function to handle file download and conversion.

    Parameters:
    - save_path: String, path to save the downloaded and processed files.
    - files: List of filenames to process.
    - number: String, used to construct the URL for downloading files.
    - gpl_path: String, path to the platform information file.

    No return value.
    """

    num = 0
    xlsx_arr = []
    for file in files:
        num, xlsx_file = write_excel(file, save_path)  # Write to Excel and return the number of rows
        xlsx_arr.append(xlsx_file)

    if num < 3:
        # Handle insufficient rows in Excel, delete empty Excel files, and try to download supplementary files from FTP
        for item in xlsx_arr:
            os.remove(item)
        try:
            if os.path.exists(suppl_path):
                shutil.copytree(suppl_path, save_path, dirs_exist_ok=True)

        except Exception as e:
            logging.error(f"Error occurred while downloading suppl files: {e}")
    else:
        # If the number of rows in the Excel meets the requirement, insert platform information and save as a txt file
        for xlsx_path in xlsx_arr:
            df = pd.read_excel(xlsx_path)
            df = insert_symbol(df, matrix_path, gpl_path)
            os.remove(xlsx_path)
            df.to_csv(f'{xlsx_path[0:-4]}txt', index=False, sep='\t')


def write_excel(file, save_path):
    """
    Extract data from the given text file and write it to an Excel table.

    Parameters:
    - file: String, path to the text file to read.
    - save_path: String, path to save the generated Excel file.

    Returns:
    - num: Integer, the number of rows written to the Excel.
    - output_path: String, the path of the generated Excel file.
    """
    num = 0
    file_name = os.path.basename(file)
    output_path = f'{save_path}/{file_name[0:-3]}xlsx'

    # Open the file, read content, and split by '!'
    with open(file, 'r', encoding='utf-8') as file:
        sections = file.read()

    sections = sections.split('!')
    sections = [section.strip() for section in sections]
    data_dict = {}  # Store extracted data

    # Parse each section and store in data_dict
    for section in sections:
        try:
            if section:
                match = re.search(r'[\t\n]', section)
                if match:
                    split_index = match.start()
                    if split_index > 0:
                        key = section[:split_index]
                        value = section[split_index + 1:]
                        data_dict[key] = value
                    else:
                        key = section
                        value = ""

        except ValueError as e:
            print(f"ValueError encountered: {e}")

    # Prepare to write to Excel
    sections = None  # Release resources
    workbook = Workbook()
    sheet = workbook.active  # Get the active worksheet

    col = []
    # Extract and process platform information
    platform = data_dict.get('Sample_platform_id', '').split('\t')
    platform = [cell.strip('"') for cell in platform]
    col.append(platform[0])

    # Extract and process protocol information
    extract_protocol = data_dict.get('Sample_title', '').split('\t')
    protocol = [cell.strip('"') for cell in extract_protocol]
    for cell in protocol:
        col.append(cell)

    # Write column headers
    col_num = 1  # Start writing data from the second row
    for item in col:
        sheet.cell(row=1, column=col_num, value=item)  # 写入数据
        col_num += 1

    # Extract and write matrix data
    matrix = data_dict.get('series_matrix_table_begin', '').split('\n')
    for row_index, row in enumerate(matrix):
        processed_row = [cell.strip('"') for cell in row.split()]

        for i, data in enumerate(processed_row):
            sheet.cell(row=row_index + 2, column=i + 1, value=data)
        num = row_index  # Update row count

    workbook.save(output_path)
    df = pd.read_excel(output_path)
    column = df.columns
    df.columns = df.iloc[0].tolist()
    df.iloc[0] = column
    df.to_excel(output_path, index=False)

    return num, output_path


def rename_file(path):
    RNA = ['RNA', 'data3', 'matrix']
    Counts = ['Counts', 'count', 'Count', 'counts', 'GSM']
    scRNA = ['scRNA', 'feature', 'barcode']
    isRNA = False
    isCounts = False
    isscRNA = False

    remove_arr = []
    for root, dirs, files in os.walk(path):
        for file in files:
            if any(char in file for char in Counts):
                isCounts = True
            elif any(char in file for char in scRNA):
                isscRNA = True
            elif any(char in file for char in RNA):
                isRNA = True
            else:
                remove_arr.append(os.path.join(root, file))

        for dir in dirs:
            if any(char in dir for char in Counts):
                isCounts = True
            elif any(char in dir for char in scRNA):
                isscRNA = True
            elif any(char in dir for char in RNA):
                isRNA = True
            else:
                for sroot, sdirs, sfiles in os.walk(os.path.join(root, dir)):
                    for item in sfiles:
                        if any(char in item for char in Counts):
                            isCounts = True
                        elif any(char in item for char in scRNA):
                            isscRNA = True
                        elif any(char in item for char in RNA):
                            isRNA = True

    for item in remove_arr:
        try:
            os.remove(item)
        except:
            pass

    if isCounts:
        shutil.move(path, f'{path}_Count')
        return f'{path}_Count'
    elif isscRNA:
        shutil.move(path, f'{path}_scRNA')
        return f'{path}_scRNA'
    elif isRNA:
        shutil.move(path, f'{path}_RNA__')
        return f'{path}_RNA__'
    else:
        shutil.move(path, f'{path}_error')


def remove_dirs(path_dir):
    """
    Attempt to remove empty directories under the specified root directory.

    """
    # Check if the directory exists and is empty
    for root, dirs, files in os.walk(path_dir):
        for d in dirs:
            if os.listdir(os.path.join(root, d)) == []:
                try:
                    os.rmdir(os.path.join(root, d))
                except Exception as e:
                    print(e)

    for root, dirs, files in os.walk(path_dir):
        for d in dirs:
            if os.listdir(os.path.join(root, d)) == []:
                try:
                    os.rmdir(os.path.join(root, d))
                except Exception as e:
                    print(e)


def formal(save_path, gpl_path, soft_path):
    if 'RNA' in os.path.basename(save_path):  # Correct RNA data type
        try:
            save_path = rna_correlation(save_path, gpl_path, soft_path)
        except Exception as e:  # Replace with actual possible exceptions
            logger.error(f"Error processing data: {e}")

    elif 'Count' in os.path.basename(save_path):

        try:
            save_path = counts_correlation(save_path, gpl_path, soft_path)
        except Exception as e:  # Replace with actual possible exceptions
            logger.error(f"Error processing data: {e}")

    if not os.listdir(save_path):
        shutil.move(save_path, f'{save_path}_443')
        return

    if (not 'error' in os.path.basename(save_path)) and (not '443' in os.path.basename(save_path)):
        for root, dirs, files in os.walk(save_path):
            for file in files:
                if file.endswith('.soft'):
                    return

    shutil.move(f'{save_path}', f'{save_path}_443')


def sum_get_data(data_file_path, gpl_path, save_path):
    d = os.path.basename(data_file_path)
    save_path = os.path.join(save_path, d)

    soft_path = os.path.join(data_file_path, 'soft')
    if not os.path.exists(soft_path):
        print(f'{d} has no soft file')
        return
    suppl_path = os.path.join(data_file_path, 'suppl')
    matrix_path = os.path.join(data_file_path, 'matrix')
    if not os.path.exists(matrix_path):
        print(f'{d} has no soft file')
        return
    if os.path.exists(save_path):
        return
    elif os.path.exists(f'{save_path}_RNA__') or os.path.exists(f'{save_path}_Count') or os.path.exists(
            f'{save_path}_scRNA') or os.path.exists(f'{save_path}_error'):
        if os.path.exists(f'{save_path}_RNA__'): shutil.rmtree(f'{save_path}_RNA__')
        if os.path.exists(f'{save_path}_Count'): shutil.rmtree(f'{save_path}_Count')
        if os.path.exists(f'{save_path}_scRNA'): shutil.rmtree(f'{save_path}_scRNA')
        if os.path.exists(f'{save_path}_error'): shutil.rmtree(f'{save_path}_error')
        os.makedirs(save_path)
        analyse_matrix(data_file_path, gpl_path, suppl_path, save_path)
        # unzip_unrar(save_path)
        save_path = rename_file(save_path)
        remove_dirs(save_path)
        formal(save_path, gpl_path, soft_path)
        print(f'{d}Operation completed')
    else:
        os.makedirs(save_path)
        analyse_matrix(data_file_path, gpl_path, suppl_path, save_path)
        # unzip_unrar(save_path)
        save_path = rename_file(save_path)
        remove_dirs(save_path)
        formal(save_path, gpl_path, soft_path)
        print(f'{d}Operation completed')


def insert_symbol(df, save_path, gpl_path):
    """
    Insert symbol and Entrez ID columns into a DataFrame.

    Based on the provided GPL path and save path, retrieve data from the GEO database,
    and insert symbol and Entrez ID information into the DataFrame.

    Parameters:
    df (pandas.DataFrame): The DataFrame to which symbol and Entrez ID will be inserted.
    save_path (str): The path where GEO files will be downloaded.
    gpl_path (str): The path where GPL data files are stored.

    Returns:
    pandas.DataFrame: The DataFrame with inserted symbol and Entrez ID columns.
    """
    if not df.columns[0]:
        raise ValueError("The first column of the DataFrame is empty.")

    first_column = df[df.columns[0]]  # Get the first column ID for matching later
    gpl_ids = df.iloc[0].to_list()[0].split(GPL_ID_SEPARATOR)  # Split to get multiple GPL IDs

    entrez_dict = {}
    symbol_dict = {}

    patterns = {
        'symbol': re.compile(re.escape('symbol'), re.IGNORECASE),
        'entrez': re.compile(re.escape('entrez'), re.IGNORECASE),
        'id': re.compile(re.escape('id'), re.IGNORECASE),
    }

    for g_id in gpl_ids:
        try:
            if os.path.exists(os.path.join(f'{gpl_path}/{g_id}/{g_id}.txt')):
                gpl_data = get_GEO(filepath=f'{gpl_path}/{g_id}/{g_id}.txt')
            else:
                gpl_data = get_GEO(g_id, destdir=f'{save_path}')  # Call third-party library to download GPL file
            if gpl_data.table.to_dict('records'):  # Non-empty
                dict_list = gpl_data.table.to_dict('records')

                for pattern_key, pattern in patterns.items():
                    matched_column = next((item for item in gpl_data.table.columns if pattern.search(item)), '')
                    if pattern_key == 'symbol':
                        symbol = matched_column
                    elif pattern_key == 'entrez':
                        entrez = matched_column
                    else:  # 'id'
                        id_column = matched_column

                entrez_dict_s = {str(d[id_column]): d[entrez] if entrez else '' for d in dict_list}
                entrez_dict_s[first_column[0]] = entrez or 'entrez_null'
                entrez_dict = {**entrez_dict, **entrez_dict_s}

                symbol_dict_s = {str(d[id_column]): d[symbol] if symbol else '' for d in dict_list}
                symbol_dict_s[first_column[0]] = symbol or 'symbol_null'
                symbol_dict = {**symbol_dict, **symbol_dict_s}
            else:
                entrez = 'entrez_null'
                symbol = 'symbol_null'
                for item in first_column:
                    entrez_dict[item] = ''
                    symbol_dict[item] = ''
        except Exception as e:
            print(f"An error occurred: {e}")

    # Insert columns into DataFrame
    def insert_column(df1, column_name, values):
        """
        Insert a column into a DataFrame.

        Parameters:
        df1 (pandas.DataFrame): The DataFrame to which the column will be inserted.
        column_name (str): The name of the new column.
        values (list): The values for the new column.

        Returns:
        pandas.DataFrame: The DataFrame with the new column inserted.
        """
        df1[column_name] = values
        cols = df1.columns.tolist()
        cols.remove(column_name)
        cols.insert(1, column_name)
        df1 = df[cols]
        return df1

    df = insert_column(df, entrez or 'entrez_null',
                       [entrez_dict[item] if item in entrez_dict else '' for item in first_column])
    df = insert_column(df, symbol or 'symbol_null',
                       [symbol_dict[item] if item in symbol_dict else '' for item in first_column])

    # Delete temporary files
    for item in gpl_ids:
        file_path = f'{save_path}/{item}.txt'
        if os.path.exists(file_path):
            os.remove(file_path)

    return df


def counts_correlation(url, gpl_path, soft_path):
    """
    Calculate and save the correlation of count data based on the provided URL and GPL path.

    Parameters:
    - url: string, the URL containing the directory of count data files.
    - gpl_path: string, the path to the GPL files (gene platform files).

    No return value.
    """
    # Extract gse_num from the URL
    gse_num = os.path.basename(url)[0:-6]
    file_catalog = []
    remove_arr = []

    # Use os.walk to traverse the directory and collect all eligible count files and files to be removed
    for root, dirs, files in os.walk(url):
        for file in files:
            if re.search('count', file, re.IGNORECASE) or re.search('GSM', file, re.IGNORECASE):
                file_catalog.append(os.path.join(root, file))
            else:
                remove_arr.append(os.path.join(root, file))

    # Initialization: Download software files and get GSM and GPL information
    get_soft_file(url, soft_path)
    gsm_arr, gpl, dict_gsm = get_gsm_gpl(url)

    # Main logic
    try:
        # Check if the number of files matches; if so, process normally; otherwise, try multiple ways to parse the files
        if (len(gsm_arr) - 1) == len(file_catalog):
            df_arr = []
            df_gsm = [[]]
            for index, item in enumerate(file_catalog):
                process_file(item, df_arr, df_gsm)

            for index, df in enumerate(df_arr):
                df = df_modify(df, df_gsm[index], gpl)
                df = insert_symbol(df, os.path.abspath('../../'), gpl_path)
                if not isinstance(df.iloc[1].tolist()[1], (int, float)):
                    df = df.drop(index=1)
                df.to_csv(f'{url}/{gse_num}_counts_{index + 1}.txt', index=False, sep='\t', encoding='utf-8')
        else:
            # If the number of files does not match, try multiple ways to parse each count file
            for index, item in enumerate(file_catalog):  # Try to read the file in xlsx, csv, or txt format
                df1 = pd.DataFrame
                if 'xlsx' in item:
                    # Process Excel files
                    df1 = pd.read_excel(item, header=0, index_col=0)
                    df1 = df1.reset_index()
                    if df1.columns[0] == 'index':
                        df1.rename(columns={df1.columns[0]: 'id'}, inplace=True)
                    break
                else:
                    # Process csv or txt files
                    for sep_s in sep:
                        df1 = pd.read_csv(item, sep=sep_s, header=0, index_col=0)
                        if len(df1.columns) > 1:
                            df1 = df1.reset_index()
                            if df1.columns[0] == 'index':
                                df1.rename(columns={df1.columns[0]: 'id'}, inplace=True)
                            break
                    if 'Unnamed' in df1.columns[2]:

                        with open(item, 'r', encoding='utf-8') as file:
                            file.readline()
                            content = file.readlines()
                            for sep_s in sep:
                                df1 = pd.read_csv(StringIO(str(content)), sep=sep_s, header=0, index_col=0)
                                if len(df1.columns) > 1:
                                    break
                        df1 = df1.reset_index()
                        if df1.columns[0] == 'index':
                            df1.rename(columns={df1.columns[0]: 'id'}, inplace=True)

                # After getting df1, extract the required content from df1
                if 'id' not in df1.columns:
                    for i, id_item in enumerate(df1.columns):  # Find the column corresponding to id
                        if re.search('id', id_item, re.IGNORECASE):
                            df1.rename(columns={df1.columns[i]: 'id'}, inplace=True)
                            break

                num_columns = len(df1.columns)
                arr = []
                char_arr = ['chr', 'start', 'end', 'strand', 'length']
                if any(c.lower() == df1.columns[2].lower() for c in char_arr) and any(
                        c.lower() == df1.columns[3].lower() for c in char_arr):
                    arr.append(df1.columns[0])
                    gsm_arr = ['ID_REF']
                    gsm_arr, arr = gsm_identify(num_columns, df1, dict_gsm, gsm_arr, arr)
                else:
                    if num_columns == len(gsm_arr):
                        for i in range(num_columns):
                            arr.append(df1.columns[i])

                    elif len(file_catalog) > 1:
                        arr.append('id')
                        gsm_arr = ['ID_REF']
                        gsm_arr, arr = gsm_identify(num_columns, df1, dict_gsm, gsm_arr, arr)

                    elif (len(file_catalog) == 1) and (num_columns > len(gsm_arr)):
                        arr.append('id')
                        for i in range(num_columns - len(gsm_arr) + 1, num_columns):
                            arr.append(df1.columns[i])

                    else:
                        raise Exception('File format error')

                df1 = df1[arr]
                df1 = df_modify(df1, gsm_arr, gpl)

                df1 = insert_symbol(df1, os.path.abspath('../../'), gpl_path)

                if ('Symbol' in df1.iloc[1].tolist()[0]) or ('symbol' in df1.iloc[1].tolist()[0]):
                    df1 = df1.drop(index=1)
                df1.to_csv(f'{url}/{gse_num}_counts{index + 1}.txt', index=False, sep='\t', encoding='utf-8')

        # Remove files that are no longer needed
        for item in file_catalog:
            safe_file_operation(item, "remove")

        for item in remove_arr:
            safe_file_operation(item, "remove", f'{url[0:-6]}')

        remove_empty_directories(url)
        shutil.move(url, f'{url[0:-6]}')
        return f'{url[0:-6]}'

    except Exception as e:
        shutil.move(url, f'{url[0:-6]}_error')
        logging.error(f"Error encountered during processing: {e}")
        return f'{url[0:-6]}_error'


def remove_empty_directories(url):
    """
    Recursively delete all empty directories under the specified path.

    :param url: String, the starting directory to traverse.
    :raises PermissionError: Raised when there is no permission to access a directory or file.
    :raises NotADirectoryError: Raised when the specified path is not a directory.
    """
    # Check if url points to a directory
    if not os.path.isdir(url):
        raise NotADirectoryError(f"{url} is not a directory.")

    try:
        for root, dirs, files in os.walk(url):
            for directory in dirs:
                dir_path = os.path.join(root, directory)

                try:
                    # List directory contents and check if it is empty
                    if not os.listdir(dir_path):
                        os.rmdir(dir_path)
                    else:
                        remove_empty_directories(dir_path)
                except PermissionError:
                    print(f"Permission denied: {dir_path}")
                except OSError as e:
                    print(f"Error removing directory {dir_path}: {e}")

    except PermissionError:
        print(f"Permission denied: {url}")
    except OSError as e:
        print(f"Error walking the directory tree at {url}: {e}")


def add_row_with_column_names(df, index0, index1):
    try:
        arr = [df.columns[0], df.columns[1]]
        df.rename(columns={df.columns[0]: index0, df.columns[1]: index1}, inplace=True)
        new_row = pd.Series(arr, index=df.columns)
        new_df = pd.DataFrame([new_row])
        df = pd.concat([new_df, df], ignore_index=True)
        df.reset_index(drop=True, inplace=True)
        return df
    except Exception as e:
        print(e)


def df_modify(df, gsm, gpl):
    df.rename(columns={df.columns[0]: gpl}, inplace=True)
    new_row = pd.Series(gsm, index=df.columns)
    new_df = pd.DataFrame([new_row])
    df = pd.concat([new_df, df], ignore_index=True)
    column = df.columns
    df.columns = df.iloc[0].tolist()
    df.iloc[0] = column
    return df


def safe_file_operation(file_path, operation, *args, **kwargs):
    """
    Safely perform file operations to avoid path traversal vulnerabilities.
    """
    # Ensure the file path is expected to avoid deleting or moving unexpected files
    if not os.path.isfile(file_path):
        logging.error(f"File operation exception: File {file_path} does not exist.")
        return
    try:
        if operation == "remove":
            os.remove(file_path)
        elif operation == "move":
            shutil.move(file_path, *args, **kwargs)
    except Exception as e:
        logging.error(f"File operation exception: {e}")


def process_file(file_path, df_arr, df_gsm):
    """
    Process the logic for a single file, including data reading and format conversion.
    """
    # This function contains the logic for processing a single file from the original code, simplified for brevity
    df1 = pd.DataFrame
    for sep_s in sep:
        df1 = pd.read_csv(file_path, sep=sep_s)
        if len(df1.columns) > 1:
            break
    if len(df1.columns) == 2:
        arr = os.path.basename(file_path).split('_', 1)
        arr[1] = arr[1].split('.count', 1)[0]
        arr[1] = arr[1].split('.txt', 1)[0]
        arr[1] = arr[1].split('txt', 1)[0]
        arr[1] = arr[1].split('count', 1)[0]
        arr[1] = arr[1].split('.tsv', 1)[0]
        arr[1] = arr[1].split('.TPM', 1)[0]
        df1 = add_row_with_column_names(df1, 'id', arr[1])
        if not df_arr:  # Check if the list storing DataFrames is empty
            df_arr.append(df1)
            df_gsm[0].append('ID_REF')
            df_gsm[0].append(arr[0])
        else:  # If not empty, merge DataFrames
            if_write = 0  # Flag to determine if writing is needed
            for df_i, df_item in enumerate(df_arr):
                if df1[df1.columns[0]].equals(
                        df_item[df_item.columns[0]]):  # Matching criterion is finding identical columns
                    df_item = pd.merge(df_item, df1, on='id')
                    df_arr[df_i] = df_item
                    df_gsm[df_i].append(arr[0])
                    if_write = 1
                    break
            if if_write == 0:  # If no match is found, create a separate DataFrame
                df_arr.append(df1)
                df_gsm.append([])
                df_gsm[len(df_gsm) - 1].append('ID_REF')
                df_gsm[len(df_gsm) - 1].append(arr[0])
    else:
        raise Exception('File format error')


def parse_file_content(data):
    sections = data.split('!')
    sections = [section.strip() for section in sections if section.strip()]  # 去除空字符串
    data_dict = {}
    for section in sections:
        try:
            key, value = section.split('=', 1)
            key = key.strip()
            value = value.strip()
            data_dict.setdefault(key, []).append(value)
        except ValueError:
            # When unable to split by '=', treat the entire section as the key with an empty string value
            data_dict.setdefault(section, []).append('')
    return data_dict


def construct_gsm_gpl(data_dict):
    gsm_arr = ['ID_REF']
    for item in data_dict['Series_sample_id']:
        gsm_arr.append(item)
    gpl = ' '.join(data_dict.get('Series_platform_id', []))  # Use .join() to optimize string concatenation

    dict_gsm = {}
    for index, item in enumerate(data_dict['Sample_title']):
        dict_gsm[item] = gsm_arr[index + 1]

    return gsm_arr, gpl, dict_gsm


def get_gsm_gpl(url):
    # Simplified validation of URL, more complex validation may be needed in actual applications
    if not os.path.isabs(url):
        print("Error: URL must be an absolute path.")
        return None, None, None

    file_data = None
    for root, dirs, files in os.walk(url):
        for file in files:
            if file.endswith('.soft'):
                try:
                    with open(os.path.join(root, file), 'r', encoding='utf-8') as soft_file:
                        file_data = soft_file.read()
                        break  # Stop traversing after finding the first matching file
                except IOError as e:
                    print(f"Error opening file: {e}")
                    return None, None, None

    if file_data is None:
        print("Error: No suitable file found.")
        return None, None, None

    try:
        data_dict = parse_file_content(file_data)
        file_data = None
        return construct_gsm_gpl(data_dict)
    except Exception as e:
        print(f"An error occurred: {e}")
        return None, None, None


def gsm_identify(num_columns, df1, dict_gsm, gsm_arr, arr):
    # Check the validity of input data
    if not df1.columns.size or df1.columns.size < num_columns:
        raise ValueError('Input DataFrame is empty or has insufficient columns')
    if not dict_gsm:
        raise ValueError('Provided dict_gsm is empty')

    # Convert dict_gsm to OrderedDict to maintain order
    dict_gsm = OrderedDict(dict_gsm)

    for i in range(num_columns - 1):
        current_column = df1.columns[i + 1]
        max_similarity = 0
        matched_key = ''
        matched_value = ''

        # Iterate through the dictionary to find the best match
        for key, value in dict_gsm.items():
            similarity = fuzz.ratio(current_column, key)
            if similarity > max_similarity:
                max_similarity = similarity
                matched_key = key
                matched_value = value

        # If a match is found and the similarity is above the threshold, add it to the result list
        if max_similarity > 50:
            arr.append(current_column)
            gsm_arr.append(matched_value)
            dict_gsm.pop(matched_key)  # Remove the matched key-value pair

    # Check the length of the result list to ensure at least 3 matches
    if len(arr) < 3:
        raise Exception('File format error: At least 3 matching columns are required')

    return gsm_arr, arr


def unzip_unrar(url):
    """
    Recursively unzip and unrar all files in the specified directory.

    Parameters:
    - url: The directory URL containing the files to be processed.

    Returns:
    None.
    """
    for root, dirs, files in os.walk(url):
        remove_arr = []  # List to store paths of files to be deleted

        # Process each file in the directory
        for file in files:
            file_path = os.path.join(root, file)  # Join the file path

            # Delete files smaller than 1KB
            if os.path.getsize(file_path) < 1024:
                remove_arr.append(file_path)

            # Process .tar files
            if file.endswith(('.tar')):
                with tarfile.open(file_path, 'r') as tar:
                    try:
                        tar.extractall(file_path[0:-4])  # Extract .tar files
                        remove_arr.append(file_path)  # Add to the list of files to be deleted
                        unzip_unrar(file_path[0:-4])  # Recursively process the extracted directory
                    except Exception as e:
                        print(f'error:{e}')  # Print the exception message

            elif file.endswith(('.gz')):
                with gzip.open(file_path, 'rb') as f_in:
                    with open(file_path[0:-3], 'wb') as f_out:
                        try:
                            shutil.copyfileobj(f_in, f_out)  # Decompress .gz files
                            remove_arr.append(file_path)  # Add to the list of files to be deleted
                            unzip_unrar(file_path[0:-3])  # Recursively process the decompressed file or directory
                        except Exception as e:
                            print(f'error:{e}')  # Print the exception message
        # Delete all files in the remove_arr list
        for item in remove_arr:
            os.remove(item)


def rna_correlation(url, gpl_path, soft_path):
    """
    Download and process RNA data files from the provided URL and gene platform path, and perform correlation analysis.

    Parameters:
    url (str): The URL where the data files are located.
    gpl_path (str): The local path to the gene platform file.
    soft_path (str): The local path to save the downloaded files.

    Returns:
    str: The processed directory path or an error path if an exception occurs.
    """
    file_arr = []
    get_soft_file(url, soft_path)
    try:
        for root, dirs, files in os.walk(url):
            for file in files:  # Get data files
                if 'data3' in file:   # Check file content

                    if file.endswith('.tsv'):  # Convert data content and add entrez, symbol
                        modify_tsv(url, file, gpl_path)

                    elif file.endswith('.xlsx'):  # Remove extra data and keep main data
                        modify_data3(url, file)

                    elif file.endswith('.txt'):
                        modify_txt(url, file, gpl_path)
                elif 'RNA' in file:
                    file_arr.append(os.path.join(root, file))

        if len(file_arr) > 0:

            modify_txt(url, file_arr, gpl_path)

        shutil.move(url, f'{url[0:-6]}')
        return f'{url[0:-6]}'

    except Exception as e:
        shutil.move(url, f'{url[0:-6]}_error')
        print(e)
        return f'{url[0:-6]}_error'


def modify_data3(url, file):
    """
    Modify and process data based on the provided URL and file name.

    Parameters:
    - url: str, the URL path of the data source.
    - file: str, the name of the file to be processed.

    Description:
    This function reads an Excel file, renames columns containing 'symbol' or 'entrez', and adds a new row with specific values at the beginning.
    Finally, it saves the processed data as a TXT file and deletes the original Excel file.
    """

    try:
        # Extract the numeric part from the URL to construct a new file name
        number = os.path.basename(url)[0:-6]
        # Construct the full file path
        path = os.path.join(url, file)
        # Read the first worksheet of the Excel file
        df = pd.read_excel(path, sheet_name=0, index_col=None)
        columns = df.columns
        # Initialize symbol and entrez column names
        ID = columns[0]
        pattern1 = 'symbol'
        pattern2 = 'entrez'
        entrez = ''
        symbol = ''

        # Regular expression matching for symbol and entrez column names
        pattern1 = re.compile(re.escape(pattern1), re.IGNORECASE)
        pattern2 = re.compile(re.escape(pattern2), re.IGNORECASE)
        for item in columns:
            if pattern1.search(item):
                symbol = item
            elif pattern2.search(item):
                entrez = item

        # If no entrez column is found, create an empty entrez column
        if entrez == '':
            entrez = 'entrez_null'
            df[entrez] = [''] * len(df)

        # If no symbol column is found, create an empty symbol column
        if symbol == '':
            symbol = 'symbol_null'
            df[symbol] = [''] * len(df)

        # Get gsm and gpl information
        gsm_arr, gpl = get_gsm_gpl(url)
        num_gsm = len(gsm_arr) - 1
        arr = [gsm_arr[0], '', '']
        for i in range(num_gsm):
            arr.append(gsm_arr[i + 1])
        gsm_arr = arr

        # Construct the new column order
        column_arr = [ID, symbol, entrez]
        for i in range(num_gsm):
            column_arr.append(columns[-(num_gsm - i)])

        # Reorder the columns
        df = df[column_arr]
        # Rename the first column to gpl
        df.rename(columns={df.columns[0]: gpl}, inplace=True)

        # Create and add a new row to the DataFrame
        new_row = pd.Series(gsm_arr, index=df.columns)
        new_df = pd.DataFrame([new_row])
        df = pd.concat([new_df, df], ignore_index=True)

        column = df.columns
        df.columns = df.iloc[0].tolist()
        df.iloc[0] = column

        # Save the processed data to a TXT file
        df.to_csv(f'{url}/{number}.txt', sep='\t', index=False, encoding='utf-8')
        os.remove(path)

    except Exception as e:
        # Log or handle the exception, such as logging an error message or notifying the user
        print(f"Error processing data: {e}")
        raise


def get_soft_file(save_path, soft_path):
    try:
        for file in os.listdir(soft_path):
            if file.endswith('.soft'):
                path = os.path.join(soft_path, file)
                save_path = os.path.join(save_path, file)
                shutil.copy(path, save_path)
                break
    except Exception as e:
        print(f"{soft_path} {e}")
        raise


def modify_tsv(url, file, gpl_path):
    """
    Modify a TSV file based on the given URL, file name, and GPL path, and insert specific rows, then save it as a new file format.

    Parameters:
    - url: str, the directory URL containing the TSV file.
    - file: str, the name of the TSV file to be processed.
    - gpl_path: str, the path to the GPL file used to obtain metadata for insertion.

    Returns:
    None
    """
    number = os.path.basename(url)[0:-6]  # Extract the number from the URL, removing the last 6 characters

    for sep_s in sep:  # Try different delimiters to read the TSV file
        try:
            path = os.path.join(url, file)  # Construct the full file path
            df = pd.read_csv(path, sep=sep_s)  # Read the TSV file using Pandas

            gsm_arr, gpl = get_gsm_gpl(url)  # Get the array to be inserted and the corresponding column name
            df.rename(columns={df.columns[0]: gpl}, inplace=True)  # Rename the first column to gpl

            new_row = pd.Series(gsm_arr, index=df.columns)  # Create a new row with values from gsm_arr and indices from df columns
            new_df = pd.DataFrame([new_row])  # Convert the new row to a DataFrame
            df = pd.concat([new_df, df], ignore_index=True)  # Insert the new row at the beginning of the DataFrame

            column = df.columns
            df.columns = df.iloc[0].tolist()
            df.iloc[0] = column

            df = insert_symbol(df, '/', gpl_path)  # Insert symbols into the data, the specific operation depends on the insert_symbol function
            df.to_csv(f'{url}/{number}_mRNA.txt', sep='\t', index=False, encoding='utf-8')  # Save the modified data as a new file format
            os.remove(path)  # Delete the TSV file
            break  # Terminate the loop if the file is successfully processed

        except Exception as e:
            print(e)


def modify_txt(url, file_arr, gpl_path):
    """
    Process files in the file array based on the provided URL and modify them according to the given GPL path.

    Parameters:
    - url: str, the base URL used for file processing and part of the output file path.
    - file_arr: list, an array of file paths to be processed.
    - gpl_path: str, the path to the GPL file used to guide data processing and modification.

    Returns:
    None
    """

    # Extract the number from the URL and initialize variables
    number = os.path.basename(url)[0:-6]
    gsm_arr, gpl, dict_gsm = get_gsm_gpl(url)
    num_gsm = len(gsm_arr)
    sep_arr = ['\t', ';', ',', ' ']

    # Check if the length of the file array matches the length of gsm_arr, and process accordingly
    if (num_gsm - 1) == len(file_arr):
        # Process the case where the lengths match, merge and modify DataFrames
        df_arr = []
        df_gsm = [[]]
        for file in file_arr:
            # Read the file, trying multiple delimiters
            df = pd.DataFrame
            for sep_s in sep_arr:
                df = pd.read_csv(file, sep=sep_s, encoding='utf-8')
                if len(df.columns) > 1:
                    break

            # Process the file name to get a specific string
            arr = os.path.basename(file).split('_', 1)
            split_arr = ['.count', '.Count', '.genes', '.result', '.txt']
            for split_item in split_arr:
                if split_item in arr[1]:
                    arr[1] = arr[1].split(split_item, 1)[0]

            # Modify and rename the DataFrame based on the number of columns
            if len(df.columns) == 2:
                df = add_row_with_column_names(df, 'id', arr[1])

            else:
                for column_name in df.columns:
                    if 'count' in column_name:
                        column_a = [df.columns[0], column_name]
                        df = df[column_a]
                        df.rename(columns={df.columns[0]: 'id', column_name: arr[1]}, inplace=True)
                        break

            # Store or merge the DataFrame
            if not df_arr:
                df_arr.append(df)
                df_gsm[0].append('ID_REF')
                df_gsm[0].append(arr[0])
            else:  # If not empty, merge the df with existing ones
                if_write = 0  # Flag to determine if it should be written
                for index, df_item in enumerate(df_arr):
                    if df[df.columns[0]].equals(df_item[df_item.columns[0]]):  # Matching criterion is finding identical columns
                        df_item = pd.merge(df_item, df, on='id')
                        df_arr[index] = df_item
                        df_gsm[index].append(arr[0])
                        if_write = 1
                        break
                if if_write == 0:  # If not found, create a separate df
                    df_arr.append(df)
                    df_gsm.append([])
                    df_gsm[len(df_gsm) - 1].append('ID_REF')
                    df_gsm[len(df_gsm) - 1].append(arr[0])

        # Further modify and save each DataFrame
        for index, df in enumerate(df_arr):
            df = df_modify(df, df_gsm[index], gpl)
            df = insert_symbol(df, os.path.abspath('/'), gpl_path)
            df.to_csv(f'{url}/{number}_counts_{index + 1}.txt', index=False, sep='\t', encoding='utf-8')

    else:
        # Process the case where the lengths do not match
        for file in file_arr:
            # Read the file, trying multiple delimiters
            df = pd.DataFrame
            if file.endswith('.xlsx'):
                df = pd.read_excel(file)
            else:
                for sep_s in sep:
                    df = pd.read_csv(file, sep=sep_s, encoding='utf-8')
                    if len(df.columns) > 1:
                        break

            # Process file data based on the number of columns and predefined rules
            num_columns = len(df.columns)
            arr = []

            if len(df.columns) == num_gsm:
                for i in range(num_columns):
                    arr.append(df.columns[i])
            else:
                arr.append(df.columns[0])
                gsm_arr = ['ID_REF']
                for i in range(num_columns - 1):
                    str1 = df.columns[i + 1]
                    radio_max = 0
                    key_max = ''
                    value_max = ''
                    for key, value in dict_gsm.items():
                        radio = fuzz.ratio(str1, key)
                        if radio > radio_max:
                            radio_max = radio
                            key_max = key
                            value_max = value
                    if radio_max > 50:
                        arr.append(str1)
                        gsm_arr.append(value_max)
                        dict_gsm.pop(key_max)
                if len(arr) < 3:
                    raise Exception('File format error')

            # Further modify and save the processed DataFrame
            df = df[arr]
            df = df_modify(df, gsm_arr, gpl)
            df = insert_symbol(df, os.path.abspath('/'), gpl_path)
            df.to_csv(f'{url}/{file[0:-4]}.txt', sep='\t', index=False, encoding='utf-8')
            os.remove(os.path.join(url, file))

    # Delete processed files and empty directories
    for item in file_arr:
        os.remove(item)
    remove_empty_directories(url)
